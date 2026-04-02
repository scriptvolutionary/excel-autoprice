from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import logging
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from seasonal_price.domain.models import AllocationLine, ImportIssue, ImportFileMeta, OrderLine, StockItem
from seasonal_price.exceptions import ValidationError
from seasonal_price.infrastructure.price_layout import (
    PriceSection,
    build_price_sections,
    format_season_title,
    split_price_sections,
)


YELLOW_FILL = PatternFill(
    fill_type="solid", start_color="FFF59D", end_color="FFF59D")
TITLE_FILL = PatternFill(fill_type="solid", start_color="FFE082", end_color="FFE082")
NOTICE_FILL = PatternFill(fill_type="solid", start_color="FFF3CD", end_color="FFF3CD")
LABEL_FILL = PatternFill(fill_type="solid", start_color="FDEBD0", end_color="FDEBD0")
SECTION_FILL = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
SECTION_NOTE_FILL = PatternFill(fill_type="solid", start_color="E2F0D9", end_color="E2F0D9")
SKU_FILL = PatternFill(fill_type="solid", start_color="E8F4FD", end_color="E8F4FD")
INPUT_FILL = PatternFill(fill_type="solid", start_color="FFFDE7", end_color="FFFDE7")
THIN_BORDER = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium", color="666666"),
    right=Side(style="medium", color="666666"),
    top=Side(style="medium", color="666666"),
    bottom=Side(style="medium", color="666666"),
)


@dataclass(frozen=True)
class StockInputRow:
    category_name: str
    sort_name: str
    container: str
    size: str
    price: float
    multiplicity: int
    stock_total: int


@dataclass(frozen=True)
class ParsedOrderRow:
    row_index: int
    sku: str | None
    sort_name: str | None
    container: str | None
    size: str | None
    quantity: int


@dataclass(frozen=True)
class ParsedOrderFile:
    meta: ImportFileMeta
    rows: list[ParsedOrderRow]
    issues: list[ImportIssue]


class ExcelProcessor:
    """Адаптер для чтения и записи Excel-файлов.

    Код использует поиск колонок по заголовкам, чтобы переживать перестановки
    столбцов в клиентских файлах.
    """

    def __init__(self, logger: logging.Logger) -> None:
        self._logger = logger

    def discover_order_files(self, input_dir: Path) -> list[Path]:
        files = [
            path
            for path in input_dir.glob("*")
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"}
        ]
        files.sort(key=lambda p: p.stat().st_mtime)
        return files

    def read_stock_sheet(self, stock_file: Path, sheet_name: str = "Склад") -> list[StockInputRow]:
        with pd.ExcelFile(stock_file) as book:
            resolved_sheet_name = self._resolve_sheet_name(
                book.sheet_names, sheet_name)
            if resolved_sheet_name is None:
                available = ", ".join(
                    f"'{name}'" for name in book.sheet_names) or "—"
                raise ValidationError(
                    f"В файле '{stock_file.name}' не найден лист '{sheet_name}'. "
                    f"Доступные листы: {available}."
                )

            df = pd.read_excel(book, sheet_name=resolved_sheet_name, dtype=object)
            header_map = self._build_header_map(df.columns)
            required = {
                "sku": None,
                "category": None,
                "sort": None,
                "container": None,
                "size": None,
                "price": None,
                "multiplicity": None,
                "stock": None,
            }
            for key, source in header_map.items():
                if key in required:
                    required[key] = source

            missing = [key for key, column in required.items() if key !=
                       "sku" and column is None]
            if missing:
                fallback_rows = self._parse_legacy_stock_sheet(
                    book=book,
                    sheet_name=resolved_sheet_name,
                )
                if fallback_rows:
                    self._logger.warning(
                        "Лист '%s' в файле '%s' не содержит стандартной шапки. Используем fallback-парсер.",
                        resolved_sheet_name,
                        stock_file.name,
                    )
                    return fallback_rows

                missing_str = ", ".join(missing)
                raise ValidationError(
                    f"Лист '{resolved_sheet_name}' не содержит обязательные колонки: {missing_str}. "
                    "Ожидаются: Категория, Сорт, Контейнер, Размер, Цена, Кратность, Остаток."
                )

            rows: list[StockInputRow] = []
            for _, row in df.iterrows():
                category = self._to_text(
                    row[required["category"]])  # type: ignore[index]
                sort_name = self._to_text(
                    row[required["sort"]])  # type: ignore[index]
                container = self._to_text(
                    row[required["container"]])  # type: ignore[index]
                size = self._to_text(row[required["size"]])  # type: ignore[index]
                if not category or not sort_name:
                    continue
                # type: ignore[index]
                price = self._to_float(row[required["price"]])
                multiplicity = self._to_int(
                    row[required["multiplicity"]])  # type: ignore[index]
                stock_total = self._to_int(
                    row[required["stock"]])  # type: ignore[index]
                rows.append(
                    StockInputRow(
                        category_name=category,
                        sort_name=sort_name,
                        container=container,
                        size=size,
                        price=price,
                        multiplicity=max(multiplicity, 1),
                        stock_total=max(stock_total, 0),
                    )
                )
            if not rows:
                raise ValidationError(
                    f"В листе '{resolved_sheet_name}' не найдено ни одной валидной товарной строки.")
            return rows

    def write_client_price(self, output_path: Path, stock_items: list[StockItem]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Прайс"
        headers = [
            "SKU",
            "Категория",
            "Культура/сорт",
            "Контейнер",
            "Размер",
            "Цена ОПТ",
            "Кратность",
            "Заказ",
            "Подтверждение",
            "Сумма",
            "Сумма по подтверждению",
            "Примечание",
        ]
        ws.append(headers)
        for item in stock_items:
            ws.append(
                [
                    item.sku,
                    item.category_name,
                    item.sort_name,
                    item.container,
                    item.size,
                    item.price,
                    item.multiplicity,
                    0,
                    0,
                    0,
                    0,
                    "",
                ]
            )

        for row_idx in range(2, ws.max_row + 1):
            ws[f"J{row_idx}"] = f"=H{row_idx}*F{row_idx}"
            ws[f"K{row_idx}"] = f"=I{row_idx}*F{row_idx}"

        for idx, header in enumerate(headers, start=1):
            width = max(len(header) + 2, 14)
            ws.column_dimensions[get_column_letter(idx)].width = width

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def read_order_file(self, file_path: Path) -> ParsedOrderFile:
        mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
        book = pd.read_excel(file_path, sheet_name=None,
                             header=None, dtype=object)
        client_name = self._extract_client_name(book, fallback=file_path.stem)

        best_rows: list[ParsedOrderRow] = []
        all_issues: list[ImportIssue] = []
        for sheet_name, frame in book.items():
            parsed_rows, issues = self._parse_order_sheet(frame, file_path)
            if len(parsed_rows) > len(best_rows):
                best_rows = parsed_rows
            all_issues.extend(issues)

        meta = ImportFileMeta(file_path=file_path,
                              client_name=client_name, mtime=mtime)
        return ParsedOrderFile(meta=meta, rows=best_rows, issues=all_issues)

    def write_import_report(
        self,
        output_path: Path,
        success_files: int,
        error_files: int,
        issues: list[ImportIssue],
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет импорта"
        ws.append(["Параметр", "Значение"])
        ws.append(["Успешно обработано файлов", success_files])
        ws.append(["Файлы с ошибками", error_files])
        ws.append([])
        ws.append(["Файл", "Код", "Сообщение", "Строка", "SKU"])

        for issue in issues:
            ws.append(
                [
                    str(issue.file_path),
                    issue.issue_code,
                    issue.message,
                    issue.row_index if issue.row_index is not None else "",
                    issue.sku if issue.sku is not None else "",
                ]
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def write_order_summary(
        self,
        output_path: Path,
        stock_items: list[StockItem],
        order_lines: list[OrderLine],
        allocation_lines: list[AllocationLine] | None = None,
    ) -> None:
        clients = sorted({line.client_name for line in order_lines})
        requested_map: dict[tuple[str, str], int] = {}
        rounded_map: dict[tuple[str, str], int] = {}
        order_changed_map: dict[tuple[str, str], bool] = {}
        for line in order_lines:
            key = (line.sku, line.client_name)
            requested_map[key] = requested_map.get(key, 0) + line.requested_qty
            rounded_map[key] = rounded_map.get(key, 0) + line.rounded_qty
            order_changed_map[key] = order_changed_map.get(
                key, False) or line.was_rounded

        conf_map: dict[tuple[str, str], int] = {}
        allocation_changed_map: dict[tuple[str, str], bool] = {}
        if allocation_lines:
            for line in allocation_lines:
                key = (line.sku, line.client_name)
                conf_map[key] = conf_map.get(key, 0) + line.confirmed_qty
                allocation_changed_map[key] = allocation_changed_map.get(
                    key, False) or line.was_rounded

        changed_map = dict(order_changed_map)
        for key, value in allocation_changed_map.items():
            changed_map[key] = changed_map.get(key, False) or value

        wb = Workbook()
        ws_req = wb.active
        ws_req.title = "Заказы"
        req_headers = [
            "SKU",
            "Категория",
            "Культура/сорт",
            "Контейнер",
            "Размер",
            "Цена",
            "Кратность",
            "Остаток",
            "Заказано",
            *clients,
        ]
        ws_req.append(req_headers)

        ws_round = wb.create_sheet("К расчету")
        round_headers = [
            "SKU",
            "Категория",
            "Культура/сорт",
            "Контейнер",
            "Размер",
            "Цена",
            "Кратность",
            "Остаток",
            "К расчету",
            *clients,
        ]
        ws_round.append(round_headers)

        ws_conf = wb.create_sheet("Подтверждения")
        conf_headers = [
            "SKU",
            "Категория",
            "Культура/сорт",
            "Контейнер",
            "Размер",
            "Цена",
            "Кратность",
            "Остаток",
            "Подтверждено",
            *clients,
        ]
        ws_conf.append(conf_headers)

        for item in stock_items:
            req_row = [
                item.sku,
                item.category_name,
                item.sort_name,
                item.container,
                item.size,
                item.price,
                item.multiplicity,
                item.stock_total,
            ]
            req_start_col = len(req_row) + 1
            req_total = 0
            req_row.append(req_total)
            for client in clients:
                qty = requested_map.get((item.sku, client), 0)
                req_row.append(qty)
                req_total += qty
            req_row[8] = req_total
            ws_req.append(req_row)
            req_row_idx = ws_req.max_row
            for offset, client in enumerate(clients):
                col_idx = req_start_col + 1 + offset
                if order_changed_map.get((item.sku, client), False):
                    ws_req.cell(row=req_row_idx,
                                column=col_idx).fill = YELLOW_FILL

            round_row = [
                item.sku,
                item.category_name,
                item.sort_name,
                item.container,
                item.size,
                item.price,
                item.multiplicity,
                item.stock_total,
            ]
            round_start_col = len(round_row) + 1
            round_total = 0
            round_row.append(round_total)
            for client in clients:
                qty = rounded_map.get((item.sku, client), 0)
                round_row.append(qty)
                round_total += qty
            round_row[8] = round_total
            ws_round.append(round_row)

            round_row_idx = ws_round.max_row
            for offset, client in enumerate(clients):
                col_idx = round_start_col + 1 + offset
                if order_changed_map.get((item.sku, client), False):
                    ws_round.cell(
                        row=round_row_idx, column=col_idx).fill = YELLOW_FILL

            conf_row = [
                item.sku,
                item.category_name,
                item.sort_name,
                item.container,
                item.size,
                item.price,
                item.multiplicity,
                item.stock_total,
            ]
            conf_total = 0
            conf_row.append(conf_total)
            start_col = len(conf_row)
            for client in clients:
                qty = conf_map.get((item.sku, client), 0)
                conf_row.append(qty)
                conf_total += qty
            conf_row[8] = conf_total
            ws_conf.append(conf_row)

            row_idx = ws_conf.max_row
            for offset, client in enumerate(clients):
                col_idx = start_col + 1 + offset
                if changed_map.get((item.sku, client), False):
                    ws_conf.cell(
                        row=row_idx, column=col_idx).fill = YELLOW_FILL

        self._auto_size(ws_req)
        self._auto_size(ws_round)
        self._auto_size(ws_conf)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def write_residual_price(self, output_path: Path, stock_items: list[StockItem]) -> None:
        self.write_client_price(output_path, stock_items)

    def write_allocation_report(self, output_path: Path, allocation_lines: list[AllocationLine]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет аллокации"
        ws.append(
            [
                "Клиент",
                "Файл",
                "Время заказа",
                "SKU",
                "Заказано",
                "К расчету",
                "Подтверждено",
                "Коррекция при аллокации",
                "Режим",
            ]
        )
        for line in sorted(allocation_lines, key=lambda x: (x.client_name, x.sku, x.order_mtime)):
            ws.append(
                [
                    line.client_name,
                    str(line.source_file),
                    line.order_mtime.isoformat(),
                    line.sku,
                    line.requested_qty,
                    line.rounded_qty,
                    line.confirmed_qty,
                    "Да" if line.was_rounded else "Нет",
                    line.allocation_mode,
                ]
            )
        self._auto_size(ws)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def write_client_confirmation_xlsx(
        self,
        output_path: Path,
        client_name: str,
        lines: list[tuple[StockItem, AllocationLine]],
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Подтверждение"
        ws.append(["Клиент", client_name])
        ws.append(["Дата", datetime.now().strftime("%Y-%m-%d %H:%M")])
        ws.append([])
        ws.append(
            ["SKU", "Категория", "Культура/сорт", "Контейнер",
                "Размер", "Цена", "Заказано", "К расчету", "Подтверждено", "Сумма"]
        )
        total = 0.0
        for stock, line in lines:
            line_sum = stock.price * line.confirmed_qty
            total += line_sum
            ws.append(
                [
                    stock.sku,
                    stock.category_name,
                    stock.sort_name,
                    stock.container,
                    stock.size,
                    stock.price,
                    line.requested_qty,
                    line.rounded_qty,
                    line.confirmed_qty,
                    line_sum,
                ]
            )
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "", "Итого", total])
        self._auto_size(ws)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    @staticmethod
    def _auto_size(worksheet: Any) -> None:
        for col_idx in range(1, worksheet.max_column + 1):
            max_len = 12
            for row_idx in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)) + 2)
            worksheet.column_dimensions[get_column_letter(
                col_idx)].width = min(max_len, 44)

    def write_client_price(self, output_path: Path, stock_items: list[StockItem]) -> None:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        sections = build_price_sections(stock_items)
        season_title = format_season_title(
            output_path.stem.replace("_Прайс_для_клиентов", "")
        )
        pages = split_price_sections(sections)
        for index, page_sections in enumerate(pages, start=1):
            title = "Прайс" if index == 1 else f"Прайс {index}"
            ws = wb.create_sheet(title=title)
            self._render_price_sheet(
                worksheet=ws,
                season_title=season_title,
                sections=page_sections,
                continuation=index > 1,
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def read_order_file(self, file_path: Path) -> ParsedOrderFile:
        mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
        book = pd.read_excel(file_path, sheet_name=None, header=None, dtype=object)
        client_name = self._extract_client_name(book, fallback=file_path.stem)

        all_rows: list[ParsedOrderRow] = []
        all_issues: list[ImportIssue] = []
        for _, frame in book.items():
            parsed_rows, issues = self._parse_order_sheet(frame, file_path)
            all_rows.extend(parsed_rows)
            all_issues.extend(issues)

        meta = ImportFileMeta(file_path=file_path, client_name=client_name, mtime=mtime)
        return ParsedOrderFile(meta=meta, rows=all_rows, issues=all_issues)

    def _render_price_sheet(
        self,
        worksheet: Any,
        season_title: str,
        sections: list[PriceSection],
        continuation: bool,
    ) -> None:
        self._setup_price_columns(worksheet)
        row_idx = self._append_price_intro(
            worksheet=worksheet,
            season_title=season_title,
            continuation=continuation,
        )
        for section in sections:
            row_idx = self._append_price_section(worksheet, row_idx, section)

        worksheet.freeze_panes = "A19"
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_title_rows = "$1:$18"
        worksheet.protection.sheet = True
        worksheet.protection.formatColumns = False
        worksheet.protection.formatRows = False
        worksheet.protection.sort = False
        worksheet.protection.autoFilter = False

    def _setup_price_columns(self, worksheet: Any) -> None:
        widths = {
            "A": 14,
            "B": 48,
            "C": 18,
            "D": 16,
            "E": 12,
            "F": 12,
            "G": 15,
            "H": 14,
            "I": 19,
            "J": 18,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    def _append_price_intro(
        self,
        worksheet: Any,
        season_title: str,
        continuation: bool,
    ) -> int:
        for row_idx in range(1, 6):
            worksheet.row_dimensions[row_idx].height = 16

        title_suffix = " (продолжение)" if continuation else ""

        worksheet.merge_cells("B6:J6")
        worksheet["B6"] = "Оптовый прайс и форма заказа. Заполните реквизиты клиента и колонку 'Заказ'."
        worksheet["B6"].font = Font(size=14)
        worksheet["B6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[6].height = 34

        worksheet.merge_cells("A8:J8")
        worksheet["A8"] = "Только оптовая продажа по предварительным заявкам."
        worksheet["A8"].fill = NOTICE_FILL
        worksheet["A8"].font = Font(size=12, bold=True)
        worksheet["A8"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[8].height = 26

        worksheet.merge_cells("A9:J9")
        worksheet["A9"] = "Не меняйте структуру файла и названия колонок. SKU обязателен для корректного импорта."
        worksheet["A9"].fill = NOTICE_FILL
        worksheet["A9"].font = Font(size=11, bold=True)
        worksheet["A9"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[9].height = 30

        worksheet.merge_cells("C10:G10")
        worksheet["C10"] = f"{season_title}{title_suffix}"
        worksheet["C10"].fill = TITLE_FILL
        worksheet["C10"].font = Font(size=16, bold=True)
        worksheet["C10"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("H10:J10")
        worksheet["H10"] = "Отправляйте клиенту этот файл без изменения структуры."
        worksheet["H10"].fill = LABEL_FILL
        worksheet["H10"].font = Font(size=10, bold=True)
        worksheet["H10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[10].height = 38

        worksheet.merge_cells("B11:J11")
        worksheet["B11"] = "ОБЯЗАТЕЛЬНО: клиент заполняет реквизиты и только колонку 'Заказ'."
        worksheet["B11"].fill = TITLE_FILL
        worksheet["B11"].font = Font(size=12, bold=True)
        worksheet["B11"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[11].height = 30

        labels = {
            12: "Фирма заказчика",
            13: "Город",
            14: "Контактное лицо",
            15: "Телефон для оперативной связи",
            16: "Email",
        }
        for row_idx, label in labels.items():
            label_cell = worksheet[f"B{row_idx}"]
            label_cell.value = label
            label_cell.fill = LABEL_FILL
            label_cell.font = Font(size=11, bold=True)
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            for col_idx in range(2, 11):
                worksheet.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            worksheet.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=10)
            data_cell = worksheet.cell(row=row_idx, column=3)
            data_cell.protection = Protection(locked=False)
            data_cell.alignment = Alignment(horizontal="left", vertical="center")
            worksheet.row_dimensions[row_idx].height = 24

        worksheet.merge_cells("B17:G17")
        worksheet["B17"] = "Заполняется только графа ЗАКАЗ"
        worksheet["B17"].font = Font(size=12, bold=True, color="9C6500")
        worksheet["B17"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet["B17"].border = Border(bottom=Side(style="thin", color="9C6500"))
        worksheet.row_dimensions[17].height = 22

        headers = [
            "SKU",
            "Культура/сорт",
            "Контейнер",
            "Размер",
            "Цена ОПТ.",
            "Заказ",
            "Подтверждение",
            "Сумма",
            "Сумма по подтверждению",
            "Примечание",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=18, column=col_idx)
            cell.value = header
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = MEDIUM_BORDER
        worksheet.row_dimensions[18].height = 42
        return 19

    def _append_price_section(self, worksheet: Any, row_idx: int, section: PriceSection) -> int:
        worksheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=10)
        for col_idx in range(1, 11):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = SECTION_FILL
            cell.border = THIN_BORDER
        title_cell = worksheet.cell(row=row_idx, column=2)
        title_cell.value = section.title
        title_cell.font = Font(size=12, bold=True, italic=True)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row_idx].height = 24
        row_idx += 1

        worksheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=10)
        for col_idx in range(1, 11):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = SECTION_NOTE_FILL
            cell.border = THIN_BORDER
        subtitle_cell = worksheet.cell(row=row_idx, column=2)
        subtitle_cell.value = section.subtitle or "Позиции к заказу"
        subtitle_cell.font = Font(size=11, italic=True)
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row_idx].height = 22
        row_idx += 1

        worksheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=10)
        for col_idx in range(1, 11):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = SECTION_NOTE_FILL
            cell.border = THIN_BORDER
        note_cell = worksheet.cell(row=row_idx, column=2)
        note_cell.value = section.multiplicity_note
        note_cell.font = Font(size=10, italic=True)
        note_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row_idx].height = 22
        row_idx += 1

        for item in section.items:
            self._append_price_item_row(worksheet, row_idx, item)
            row_idx += 1

        for col_idx in range(1, 11):
            worksheet.cell(row=row_idx, column=col_idx).border = Border(
                top=Side(style="thin", color="D0D0D0")
            )
        worksheet.row_dimensions[row_idx].height = 8
        return row_idx + 1

    def _append_price_item_row(self, worksheet: Any, row_idx: int, item: StockItem) -> None:
        values = [
            item.sku,
            item.sort_name,
            item.container,
            item.size,
            item.price,
            "",
            "",
            f"=IFERROR(F{row_idx}*E{row_idx},0)",
            f"=IFERROR(G{row_idx}*E{row_idx},0)",
            "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in {2, 10}))
            if col_idx == 1:
                cell.fill = SKU_FILL
                cell.font = Font(size=10, bold=True)
            elif col_idx == 2:
                cell.fill = YELLOW_FILL
                cell.font = Font(size=11)
            elif col_idx in {6, 10}:
                cell.fill = INPUT_FILL
                cell.protection = Protection(locked=False)
                cell.font = Font(size=11)
            elif col_idx == 5:
                cell.font = Font(size=11, bold=True)
                cell.number_format = "#,##0"
            elif col_idx in {8, 9}:
                cell.number_format = "#,##0"
                cell.font = Font(size=11)
            else:
                cell.font = Font(size=11)

        worksheet.cell(row=row_idx, column=6).number_format = "0"
        worksheet.cell(row=row_idx, column=7).number_format = "0"
        worksheet.cell(row=row_idx, column=8).number_format = "#,##0"
        worksheet.cell(row=row_idx, column=9).number_format = "#,##0"
        worksheet.row_dimensions[row_idx].height = 22

    def _extract_client_name(self, sheets: dict[str, pd.DataFrame], fallback: str) -> str:
        needle = "фирма заказчика"
        for frame in sheets.values():
            matrix = frame.where(pd.notna(frame), "").astype(str)
            for row_idx in range(matrix.shape[0]):
                row = [str(cell).strip()
                       for cell in matrix.iloc[row_idx].tolist()]
                for col_idx, cell in enumerate(row):
                    if needle in cell.casefold():
                        for offset in range(1, 6):
                            idx = col_idx + offset
                            if idx < len(row):
                                value = row[idx].strip()
                                if value:
                                    return value
        return fallback

    def _parse_order_sheet(
        self, frame: pd.DataFrame, file_path: Path
    ) -> tuple[list[ParsedOrderRow], list[ImportIssue]]:
        issues: list[ImportIssue] = []
        matrix = frame.where(pd.notna(frame), "").astype(str)
        header_pos = self._find_order_header_row(matrix)
        if header_pos is None:
            return [], issues

        row_idx, column_map = header_pos
        rows: list[ParsedOrderRow] = []
        for idx in range(row_idx + 1, matrix.shape[0]):
            line = [str(cell).strip() for cell in matrix.iloc[idx].tolist()]
            qty_raw = line[column_map["qty"]] if "qty" in column_map else ""
            if not qty_raw:
                continue

            qty = self._parse_quantity(qty_raw)
            if qty is None:
                issues.append(
                    ImportIssue(
                        file_path=file_path,
                        issue_code="invalid_quantity",
                        message=f"Некорректное количество '{qty_raw}'. Ожидается целое >= 0.",
                        row_index=idx + 1,
                    )
                )
                continue
            sku = None
            if "sku" in column_map:
                raw_sku = line[column_map["sku"]].strip()
                sku = raw_sku if raw_sku else None

            sort_name = line[column_map["sort"]].strip(
            ) if "sort" in column_map else None
            container = line[column_map["container"]].strip(
            ) if "container" in column_map else None
            size = line[column_map["size"]].strip(
            ) if "size" in column_map else None

            if not sku and not sort_name:
                issues.append(
                    ImportIssue(
                        file_path=file_path,
                        issue_code="missing_identity",
                        message="Строка заказа не содержит ни SKU, ни названия сорта.",
                        row_index=idx + 1,
                    )
                )
                continue
            rows.append(
                ParsedOrderRow(
                    row_index=idx + 1,
                    sku=sku,
                    sort_name=sort_name,
                    container=container,
                    size=size,
                    quantity=qty,
                )
            )
        return rows, issues

    def _find_order_header_row(
        self, matrix: pd.DataFrame
    ) -> tuple[int, dict[str, int]] | None:
        synonyms = {
            "sku": ("sku", "артикул"),
            "sort": ("культура/сорт", "сорт", "позиция"),
            "container": ("контейнер",),
            "size": ("размер",),
            "qty": ("заказ", "количество"),
        }
        for row_idx in range(matrix.shape[0]):
            row = [str(cell).strip().casefold()
                   for cell in matrix.iloc[row_idx].tolist()]
            column_map: dict[str, int] = {}
            for col_idx, value in enumerate(row):
                for key, names in synonyms.items():
                    if any(name in value for name in names):
                        column_map[key] = col_idx
            if "qty" in column_map and ("sku" in column_map or "sort" in column_map):
                return row_idx, column_map
        return None

    @staticmethod
    def _parse_quantity(raw_value: str) -> int | None:
        candidate = raw_value.replace(" ", "").replace(",", ".")
        if not candidate:
            return None
        if re.fullmatch(r"\d+", candidate):
            return int(candidate)
        if re.fullmatch(r"\d+\.0+", candidate):
            return int(float(candidate))
        return None

    @staticmethod
    def _resolve_sheet_name(sheet_names: list[str], expected_name: str) -> str | None:
        for name in sheet_names:
            if name == expected_name:
                return name

        expected_normalized = expected_name.strip().casefold()
        for name in sheet_names:
            if str(name).strip().casefold() == expected_normalized:
                return name
        return None

    def _parse_legacy_stock_sheet(self, book: pd.ExcelFile, sheet_name: str) -> list[StockInputRow]:
        frame = pd.read_excel(book, sheet_name=sheet_name,
                              header=None, dtype=object)
        matrix = frame.where(pd.notna(frame), "")
        rows: list[StockInputRow] = []
        current_category = ""
        current_multiplicity = 1

        for _, source_row in matrix.iterrows():
            cells = [self._to_text(value) for value in source_row.tolist()]
            if not any(cells):
                continue

            combined = " ".join(cells).casefold()
            if "заказ крат" in combined:
                detected = self._extract_pack_size(combined)
                if detected is not None:
                    current_multiplicity = max(detected, 1)

            category_candidate = self._extract_legacy_category(cells)
            if category_candidate is not None:
                current_category = category_candidate

            item_no = cells[1] if len(cells) > 1 else ""
            if not self._is_legacy_item_number(item_no):
                continue

            sort_name = cells[2] if len(cells) > 2 else ""
            if not sort_name:
                continue
            container = cells[3] if len(cells) > 3 else ""
            size = cells[4] if len(cells) > 4 else ""
            price_raw = cells[5] if len(cells) > 5 else ""
            stock_raw = cells[6] if len(cells) > 6 else ""
            pack_size = self._extract_pack_size(
                container) or current_multiplicity
            rows.append(
                StockInputRow(
                    category_name=current_category or "Без категории",
                    sort_name=sort_name,
                    container=container,
                    size=size,
                    price=self._to_float(price_raw),
                    multiplicity=max(pack_size, 1),
                    stock_total=max(self._to_int(stock_raw), 0),
                )
            )
        return rows

    @staticmethod
    def _extract_legacy_category(cells: list[str]) -> str | None:
        if len(cells) < 2:
            return None
        value = cells[1].strip()
        if not value:
            return None
        normalized = value.casefold()
        if "заказ крат" in normalized:
            return None
        if re.fullmatch(r"\d+", value.replace(" ", "")):
            return None
        return value

    @staticmethod
    def _extract_pack_size(text: str) -> int | None:
        match = re.search(r"(\d+)\s*(?:шт|штук|штуки)", text.casefold())
        if match is None:
            return None
        return int(match.group(1))

    @staticmethod
    def _is_legacy_item_number(value: str) -> bool:
        return re.fullmatch(r"\d+", value.replace(" ", "")) is not None

    @staticmethod
    def _build_header_map(columns: Any) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for col in columns:
            norm = str(col).strip().casefold()
            if "sku" in norm or "артикул" in norm:
                mapping["sku"] = col
            elif "категор" in norm:
                mapping["category"] = col
            elif "сорт" in norm or "культура" in norm:
                mapping["sort"] = col
            elif "контейнер" in norm:
                mapping["container"] = col
            elif "размер" in norm:
                mapping["size"] = col
            elif "цен" in norm:
                mapping["price"] = col
            elif "крат" in norm:
                mapping["multiplicity"] = col
            elif "остат" in norm:
                mapping["stock"] = col
        return mapping

    @staticmethod
    def _to_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return "" if text.lower() == "nan" else text

    @staticmethod
    def _to_int(value: Any) -> int:
        if value is None or str(value).strip() == "":
            return 0
        try:
            return int(float(str(value).replace(",", ".")))
        except ValueError as exc:
            raise ValidationError(
                f"Не удалось преобразовать значение '{value}' в целое число.") from exc

    @staticmethod
    def _to_float(value: Any) -> float:
        if value is None or str(value).strip() == "":
            return 0.0
        try:
            return float(str(value).replace(",", "."))
        except ValueError as exc:
            raise ValidationError(
                f"Не удалось преобразовать значение '{value}' в число.") from exc
