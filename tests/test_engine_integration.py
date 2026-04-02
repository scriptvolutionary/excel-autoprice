from __future__ import annotations

import os
from pathlib import Path
import shutil
from unittest.mock import patch
from uuid import uuid4
import unittest

from openpyxl import Workbook, load_workbook

from seasonal_price.application.api import (
    allocate,
    build_residual_price,
    close_season,
    export_confirmations,
    generate_price,
    import_orders,
    init_season,
)
from seasonal_price.exceptions import ValidationError


class EngineIntegrationTests(unittest.TestCase):
    def test_generate_price_accepts_case_insensitive_stock_sheet_name(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            stock_file = base_dir / "stock_case.xlsx"
            self._create_stock_file(stock_file, sheet_name="склад")

            result = generate_price(
                stock_file=stock_file,
                season_id="spring_2026",
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            self.assertTrue(Path(result["output_file"]).exists())
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_generate_price_reports_available_sheet_names(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            stock_file = base_dir / "stock_missing_sheet.xlsx"
            self._create_stock_file(stock_file, sheet_name="Ост")

            with self.assertRaises(ValidationError) as exc:
                generate_price(
                    stock_file=stock_file,
                    season_id="spring_2026",
                    output_dir=base_dir / "out",
                    base_dir=base_dir,
                )

            message = str(exc.exception)
            self.assertIn("не найден лист 'Склад'", message)
            self.assertIn("'Ост'", message)
            self.assertIn(stock_file.name, message)
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_generate_price_supports_legacy_stock_layout_without_headers(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            stock_file = base_dir / "stock_legacy.xlsx"
            self._create_legacy_stock_file(stock_file, sheet_name="склад")

            result = generate_price(
                stock_file=stock_file,
                season_id="spring_2026",
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            self.assertEqual(result["items"], 2)
            self.assertTrue(Path(result["output_file"]).exists())
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_import_orders_works_when_stderr_is_unavailable(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            init_season(season_id=season_id, base_dir=base_dir)
            price_result = generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            sku_map = self._read_generated_skus(
                Path(price_result["output_file"]))

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            self._create_order_file(
                orders_dir / "order_client_1.xlsx",
                client="Клиент 1",
                sku=sku_map["Анабель"],
                sort_name="Анабель",
                qty=30,
            )

            with patch("seasonal_price.application.engine.sys.stderr", None):
                import_result = import_orders(
                    input_dir=orders_dir,
                    season_id=season_id,
                    profile_id=profile_id,
                    base_dir=base_dir,
                    duplicate_strategy="latest",
                )

            self.assertEqual(import_result["error_files"], 0)
            self.assertEqual(import_result["success_files"], 1)
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_import_reports_missing_identity_rows(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            init_season(season_id=season_id, base_dir=base_dir)
            generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            self._create_order_file_with_rows(
                orders_dir / "order_invalid.xlsx",
                client="Клиент 1",
                rows=[{"sku": "", "sort_name": "", "container": "40 шт", "size": "2-3 листа", "qty": 12}],
            )

            import_result = import_orders(
                input_dir=orders_dir,
                season_id=season_id,
                profile_id=profile_id,
                base_dir=base_dir,
                duplicate_strategy="latest",
            )

            report_wb = load_workbook(import_result["report_file"])
            try:
                ws = report_wb["Отчет импорта"]
                rows = list(ws.iter_rows(min_row=6, values_only=True))
            finally:
                report_wb.close()

            self.assertTrue(any(row[1] == "missing_identity" for row in rows))
            self.assertEqual(import_result["success_files"], 0)
            self.assertEqual(import_result["error_files"], 1)
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_import_skips_issues_from_non_selected_duplicate_files(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            price_result = generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            sku_map = self._read_generated_skus(Path(price_result["output_file"]))

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            older_path = orders_dir / "client_old.xlsx"
            newer_path = orders_dir / "client_new.xlsx"
            self._create_order_file_with_rows(
                older_path,
                client="Клиент 1",
                rows=[{"sku": sku_map["Анабель"], "sort_name": "Анабель", "container": "40 шт", "size": "2-3 листа", "qty": "abc"}],
            )
            self._create_order_file_with_rows(
                newer_path,
                client="Клиент 1",
                rows=[{"sku": sku_map["Анабель"], "sort_name": "Анабель", "container": "40 шт", "size": "2-3 листа", "qty": 24}],
            )
            os.utime(older_path, (1_700_000_000, 1_700_000_000))
            os.utime(newer_path, (1_700_000_100, 1_700_000_100))

            import_result = import_orders(
                input_dir=orders_dir,
                season_id=season_id,
                profile_id=profile_id,
                base_dir=base_dir,
                duplicate_strategy="latest",
            )

            report_wb = load_workbook(import_result["report_file"])
            try:
                ws = report_wb["Отчет импорта"]
                rows = list(ws.iter_rows(min_row=6, values_only=True))
            finally:
                report_wb.close()

            older_matches = [
                row for row in rows if row[0] and str(older_path) in str(row[0])]
            self.assertEqual(older_matches, [])
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_legacy_match_is_logged_and_reports_separate_quantities(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            init_season(season_id=season_id, base_dir=base_dir)
            generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            self._create_order_file(
                orders_dir / "order_client_legacy.xlsx",
                client="Клиент 1",
                sku="",
                sort_name="Анабель",
                qty=30,
            )

            with self.assertLogs("seasonal_price", level="WARNING") as captured:
                import_result = import_orders(
                    input_dir=orders_dir,
                    season_id=season_id,
                    profile_id=profile_id,
                    base_dir=base_dir,
                    duplicate_strategy="latest",
                )

            self.assertTrue(
                any("legacy-сопоставление без SKU" in message for message in captured.output)
            )
            self.assertTrue(Path(import_result["summary_file"]).exists())

            alloc_result = allocate(
                season_id=season_id,
                profile_id=profile_id,
                mode="fifo",
                base_dir=base_dir,
            )

            summary_wb = load_workbook(alloc_result["summary_file"])
            try:
                ws_req = summary_wb["Заказы"]
                ws_round = summary_wb["К расчету"]
                ws_conf = summary_wb["Подтверждения"]
                self.assertEqual(ws_req["I1"].value, "Заказано")
                self.assertEqual(ws_round["I1"].value, "К расчету")
                self.assertEqual(ws_conf["I1"].value, "Подтверждено")
                self.assertEqual(ws_req["I2"].value, 30)
                self.assertEqual(ws_round["I2"].value, 24)
                self.assertEqual(ws_conf["I2"].value, 24)
                self.assertEqual(ws_req["J2"].value, 30)
                self.assertEqual(ws_round["J2"].value, 24)
                self.assertEqual(ws_conf["J2"].value, 24)
            finally:
                summary_wb.close()

            conf_result = export_confirmations(
                season_id=season_id,
                output_dir=base_dir / "confirmations",
                pdf_mode="builtin",
                base_dir=base_dir,
            )
            confirmation_dir = Path(conf_result["output_dir"])
            confirmation_file = next(confirmation_dir.rglob("*.xlsx"))
            confirmation_wb = load_workbook(confirmation_file)
            try:
                ws = confirmation_wb["Подтверждение"]
                self.assertEqual(ws["G5"].value, 30)
                self.assertEqual(ws["H5"].value, 24)
                self.assertEqual(ws["I5"].value, 24)
            finally:
                confirmation_wb.close()
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_import_orders_supports_round_up_mode(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            price_result = generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            sku_map = self._read_generated_skus(Path(price_result["output_file"]))

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            self._create_order_file(
                orders_dir / "order_up.xlsx",
                client="Клиент 1",
                sku=sku_map["Анабель"],
                sort_name="Анабель",
                qty=30,
            )

            import_result = import_orders(
                input_dir=orders_dir,
                season_id=season_id,
                profile_id=profile_id,
                base_dir=base_dir,
                duplicate_strategy="latest",
                rounding_mode="up",
            )

            self.assertEqual(import_result["rounding_mode"], "up")
            summary_wb = load_workbook(import_result["summary_file"])
            try:
                ws_round = summary_wb["К расчету"]
                self.assertEqual(ws_round["I2"].value, 48)
                self.assertEqual(ws_round["J2"].value, 48)
            finally:
                summary_wb.close()
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_confirmations_include_positions_without_stock(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            price_result = generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            sku_map = self._read_generated_skus(Path(price_result["output_file"]))

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            first_path = orders_dir / "first.xlsx"
            second_path = orders_dir / "second.xlsx"
            self._create_order_file(first_path, "Клиент 1", sku_map["Анабель"], "Анабель", 96)
            self._create_order_file(second_path, "Клиент 2", sku_map["Анабель"], "Анабель", 96)
            os.utime(first_path, (1_700_000_000, 1_700_000_000))
            os.utime(second_path, (1_700_000_100, 1_700_000_100))

            import_orders(
                input_dir=orders_dir,
                season_id=season_id,
                profile_id=profile_id,
                base_dir=base_dir,
                duplicate_strategy="latest",
            )
            allocate(
                season_id=season_id,
                profile_id=profile_id,
                mode="fifo",
                base_dir=base_dir,
            )

            conf_result = export_confirmations(
                season_id=season_id,
                output_dir=base_dir / "confirmations",
                pdf_mode="builtin",
                base_dir=base_dir,
            )

            self.assertEqual(conf_result["clients"], 2)
            confirmation_dir = Path(conf_result["output_dir"])
            second_confirmation = next(confirmation_dir.rglob("Клиент 2.xlsx"))
            confirmation_wb = load_workbook(second_confirmation)
            try:
                ws = confirmation_wb["Подтверждение"]
                self.assertEqual(ws["G5"].value, 96)
                self.assertEqual(ws["H5"].value, 96)
                self.assertEqual(ws["I5"].value, 0)
            finally:
                confirmation_wb.close()
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_end_to_end_flow(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)

            init_season(season_id=season_id, base_dir=base_dir)
            price_result = generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            self.assertTrue(Path(price_result["output_file"]).exists())

            sku_map = self._read_generated_skus(
                Path(price_result["output_file"]))
            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            self._create_order_file(
                orders_dir / "order_client_1.xlsx",
                client="Клиент 1",
                sku=sku_map["Анабель"],
                sort_name="Анабель",
                qty=30,
            )
            self._create_order_file(
                orders_dir / "order_client_2.xlsx",
                client="Клиент 2",
                sku="",
                sort_name="Блю",
                qty=48,
            )

            import_result = import_orders(
                input_dir=orders_dir,
                season_id=season_id,
                profile_id=profile_id,
                base_dir=base_dir,
                duplicate_strategy="latest",
            )
            self.assertEqual(import_result["error_files"], 0)
            self.assertEqual(import_result["success_files"], 2)
            self.assertTrue(Path(import_result["summary_file"]).exists())

            alloc_result = allocate(
                season_id=season_id,
                profile_id=profile_id,
                mode="fifo",
                base_dir=base_dir,
            )
            self.assertGreater(alloc_result["allocation_lines"], 0)
            self.assertTrue(Path(alloc_result["allocation_report"]).exists())

            conf_result = export_confirmations(
                season_id=season_id,
                output_dir=base_dir / "confirmations",
                pdf_mode="builtin",
                base_dir=base_dir,
            )
            self.assertGreater(conf_result["clients"], 0)
            self.assertTrue(Path(conf_result["output_dir"]).exists())

            residual = build_residual_price(
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            self.assertTrue(Path(residual["output_file"]).exists())

            close = close_season(
                season_id=season_id,
                archive_dir=base_dir / "archive",
                base_dir=base_dir,
            )
            self.assertTrue(Path(close["archive_dir"]).exists())
            self.assertTrue(str(close["new_season"]).startswith("season_"))
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    @staticmethod
    def _create_stock_file(path: Path, sheet_name: str = "Склад") -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Проверяем устойчивость к перестановке колонок через заголовки.
        ws.append(["Размер", "Категория", "Сорт", "Остаток",
                  "Контейнер", "Кратность", "Цена"])
        ws.append(["2-3 листа", "Гортензия", "Анабель", 100, "40 шт", 24, 42])
        ws.append(["2-3 листа", "Гортензия", "Блю", 50, "40 шт", 24, 45])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def _create_legacy_stock_file(path: Path, sheet_name: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["", "", "Закрытая корневая система", "", "", "", ""])
        ws.append(["", "Земляника садовая", "", "", "", "", ""])
        ws.append(["", "Заказ кратен 40 (кассета 40 шт)", "", "", "", "", ""])
        ws.append(["", "1", "Анабель", "40 шт", "2-3 листа", 42, 5000])
        ws.append(["", "2", "Блю", "40 шт", "2-3 листа", 45, 1500])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def _create_order_file(path: Path, client: str, sku: str, sort_name: str, qty: int) -> None:
        EngineIntegrationTests._create_order_file_with_rows(
            path,
            client=client,
            rows=[
                {
                    "sku": sku,
                    "sort_name": sort_name,
                    "container": "40 шт",
                    "size": "2-3 листа",
                    "qty": qty,
                }
            ],
        )

    @staticmethod
    def _create_order_file_with_rows(path: Path, client: str, rows: list[dict[str, object]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказ"
        ws.append(["Фирма заказчика", client])
        ws.append([])
        ws.append(["SKU", "Культура/сорт", "Контейнер", "Размер", "Заказ"])
        for row in rows:
            ws.append(
                [
                    row.get("sku", ""),
                    row.get("sort_name", ""),
                    row.get("container", "40 шт"),
                    row.get("size", "2-3 листа"),
                    row.get("qty", 0),
                ]
            )
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def _read_generated_skus(price_file: Path) -> dict[str, str]:
        wb = load_workbook(price_file)
        try:
            ws = wb["Прайс"]
            out: dict[str, str] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                sku = str(row[0])
                sort_name = str(row[2])
                out[sort_name] = sku
            return out
        finally:
            wb.close()
    def test_generate_price_creates_structured_workbook_and_pdf(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)

            result = generate_price(
                stock_file=stock_file,
                season_id="spring_2026",
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )

            self.assertTrue(Path(result["output_file"]).exists())
            self.assertTrue(Path(result["pdf_file"]).exists())

            wb = load_workbook(result["output_file"])
            try:
                ws = wb["Прайс"]
                self.assertEqual(ws["A18"].value, "SKU")
                self.assertEqual(ws["B18"].value, "Культура/сорт")
                self.assertTrue(ws.protection.sheet)
                self.assertIn("B6:J6", {str(item) for item in ws.merged_cells.ranges})
            finally:
                wb.close()
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    def test_import_orders_aggregates_rows_from_multiple_sheets(self) -> None:
        tmp_root = Path.cwd() / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        base_dir = tmp_root / f"seasonal_price_{uuid4().hex}"
        base_dir.mkdir(parents=True, exist_ok=True)
        try:
            season_id = "spring_2026"
            profile_id = "default"

            stock_file = base_dir / "stock.xlsx"
            self._create_stock_file(stock_file)
            price_result = generate_price(
                stock_file=stock_file,
                season_id=season_id,
                output_dir=base_dir / "out",
                base_dir=base_dir,
            )
            sku_map = self._read_generated_skus(Path(price_result["output_file"]))

            orders_dir = base_dir / "orders"
            orders_dir.mkdir(parents=True, exist_ok=True)
            self._create_multi_sheet_order_file(
                orders_dir / "order_multi.xlsx",
                client="Клиент 1",
                sheets=[
                    [
                        {
                            "sku": sku_map["Анабель"],
                            "sort_name": "Анабель",
                            "container": "40 шт",
                            "size": "2-3 листа",
                            "qty": 24,
                        }
                    ],
                    [
                        {
                            "sku": sku_map["Блю"],
                            "sort_name": "Блю",
                            "container": "40 шт",
                            "size": "2-3 листа",
                            "qty": 24,
                        }
                    ],
                ],
            )

            import_result = import_orders(
                input_dir=orders_dir,
                season_id=season_id,
                profile_id=profile_id,
                base_dir=base_dir,
                duplicate_strategy="latest",
            )

            summary_wb = load_workbook(import_result["summary_file"])
            try:
                ws = summary_wb["Заказы"]
                rows = {
                    str(row[2]): row[9]
                    for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[2]
                }
            finally:
                summary_wb.close()

            self.assertEqual(import_result["success_files"], 1)
            self.assertEqual(rows["Анабель"], 24)
            self.assertEqual(rows["Блю"], 24)
        finally:
            shutil.rmtree(base_dir, ignore_errors=True)

    @staticmethod
    def _create_multi_sheet_order_file(
        path: Path,
        client: str,
        sheets: list[list[dict[str, object]]],
    ) -> None:
        wb = Workbook()
        for index, rows in enumerate(sheets):
            ws = wb.active if index == 0 else wb.create_sheet()
            ws.title = "Заказ" if index == 0 else f"Заказ {index + 1}"
            ws.append(["Фирма заказчика", client])
            ws.append([])
            ws.append(["SKU", "Культура/сорт", "Контейнер", "Размер", "Заказ"])
            for row in rows:
                ws.append(
                    [
                        row.get("sku", ""),
                        row.get("sort_name", ""),
                        row.get("container", "40 шт"),
                        row.get("size", "2-3 листа"),
                        row.get("qty", 0),
                    ]
                )
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def _read_generated_skus(price_file: Path) -> dict[str, str]:
        wb = load_workbook(price_file)
        try:
            out: dict[str, str] = {}
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=19, values_only=True):
                    sku = row[0]
                    sort_name = row[1]
                    if not isinstance(sku, str) or "-" not in sku:
                        continue
                    if not isinstance(sort_name, str) or not sort_name.strip():
                        continue
                    out[sort_name.strip()] = sku.strip()
            return out
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
